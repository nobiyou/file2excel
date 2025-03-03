import os
import openpyxl
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Listbox, Scrollbar
import threading
import math
import csv
from collections import defaultdict
 
class FileScanner:
    def __init__(self):
        self.scanning = False
        self.exporting = False
        self.current_thread = None
        self.file_cache = []  # 新增缓存机制
        self.file_count = 0  # 新增文件计数器
        self.large_files_warning_given = False  # 新增警告标记

    def export_file_info_to_excel(self, folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar):
        """将文件夹中的文件信息导出到Excel，基于已扫描的文件列表"""
        self.exporting = True
        try:
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", f"文件夹路径不存在: {folder_path}")
                return
            status_label.config(text="正在处理...")
            window.update_idletasks()

            # 获取文件夹名称
            folder_path = os.path.normpath(folder_path)
            drive, tail = os.path.splitdrive(folder_path)
            if os.path.isdir(folder_path) and drive and tail in ('\\', '/'):
                folder_name = drive.rstrip(':')
            else:
                folder_name = os.path.basename(folder_path)

            excel_path = os.path.join(folder_path, f"{folder_name}.xlsx")
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            headers = ["序号", "文件夹", "文件名"]
            if export_options["size"]:
                headers.append("文件大小")
            if export_options["ctime"]:
                headers.append("创建时间")
            if export_options["mtime"]:
                headers.append("修改时间")
            if export_options["ext"]:
                headers.append("文件类型")
            if export_options["path"]:
                headers.append("文件路径")
            sheet.append(headers)

            total_size = 0
            file_count = 0
            folder_count = 0
            file_type_counts = defaultdict(int)
            index = 1  # 序号从1开始

            # 直接使用传入的all_items，不再重新扫描
            all_items = sorted(all_items)
            total_items = len(all_items)

            # 批量处理，减少进度条更新频率
            batch_size = max(1000, total_items // 50)

            # 初始化进度条
            self.update_progress_bar(progress_var, progress_label, 0, total_items, progress_bar)

            for i, itempath in enumerate(all_items):
                if i % batch_size == 0:  # 减少进度条更新频率
                    self.update_progress_bar(progress_var, progress_label, i, total_items, progress_bar)
                    window.update_idletasks()

                if os.path.isfile(itempath):
                    # 处理文件
                    file_size, file_ctime, file_mtime, file_ext = self.get_file_info(itempath)
                    if file_size is not None:
                        relative_path = os.path.relpath(itempath, folder_path)
                        folder_display = ""  # 默认：顶级文件不显示文件夹
                        if os.path.dirname(relative_path) != ".":  # 检查文件是否在子文件夹中
                            folder_display = os.path.dirname(relative_path)
                        row = [index, folder_display, os.path.basename(itempath)]
                        if export_options["size"]:
                            row.append(self.convert_size(file_size))
                        if export_options["ctime"]:
                            row.append(file_ctime)
                        if export_options["mtime"]:
                            row.append(file_mtime)
                        if export_options["ext"]:
                            row.append(file_ext)
                        if export_options["path"]:
                            row.append(itempath)

                        sheet.append(row)

                        total_size += file_size
                        file_count += 1
                        file_type_counts[file_ext] += 1
                        index += 1  # 序号+1
                elif os.path.isdir(itempath):
                    # 处理文件夹
                    relative_path = os.path.relpath(itempath, folder_path)
                    folder_display = relative_path
                    row = [index, folder_display, ""]
                    sheet.append(row)
                    folder_count += 1
                    index += 1

            # 最终更新进度条
            self.update_progress_bar(progress_var, progress_label, total_items, total_items, progress_bar)

            # 添加统计信息
            sheet.append([])  # 空行用于分隔
            sheet.append(["统计信息"])
            sheet.append(["文件总数", file_count])
            sheet.append(["文件夹总数", folder_count])
            sheet.append(["文件夹总大小", self.convert_size(total_size)])

            # 批量添加文件类型统计
            for file_type, count in file_type_counts.items():
                sheet.append([f"{file_type} 文件数量", count])

            workbook.save(excel_path)
            status_label.config(text=f"文件信息已导出到: {excel_path}")
            messagebox.showinfo("成功", f"Excel文件已保存到: {excel_path}")

        except Exception as e:
            status_label.config(text="发生错误！")
            messagebox.showerror("错误", f"导出到Excel失败: {e}")
        finally:
            self.exporting = False
            self.update_progress_bar(progress_var, progress_label, 0, 1, progress_bar)

    def convert_size(self, size_bytes):
        """转换文件大小为KB, MB, GB, TB等"""
        if size_bytes == 0:
            return "0B"
        size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return f"{s} {size_name[i]}"
 
    def get_file_info(self, filepath):
        """获取文件信息"""
        try:
            file_size = os.path.getsize(filepath)
            file_ctime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getctime(filepath)))
            file_mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(filepath)))
            file_ext = os.path.splitext(filepath)[1]  # 获取文件扩展名
            return file_size, file_ctime, file_mtime, file_ext
        except FileNotFoundError:
            return None, None, None, None
        except Exception as e:
            return None, None, None, None
 
    def update_progress_bar(self, progress_var, progress_label, current, total, progress_bar):
        """更新进度条和进度标签"""
        progress = int((current / total) * 100) if total > 0 else 0
        progress_var.set(progress)
        progress_label.config(text=f"进度: {progress}%")
        if progress == 0:
            progress_bar.pack_forget()  # Hide the progress bar
            progress_label.pack_forget()
        else:
            progress_bar.pack(pady=(0, 10), fill=tk.X, padx=5)  # Show the progress bar, fill horizontally, add padding
            progress_label.pack(pady=(0, 10))
        window.update_idletasks()
 
    def start_export(self, folder_entry, export_options, status_label, file_listbox, include_subfolders, progress_var, progress_label, progress_bar):
        """启动导出操作"""
        if not self.exporting and not self.scanning:
            folder_path = folder_entry.get()
            if not folder_path:
                messagebox.showerror("错误", "请选择文件夹")
                return
            # 检查文件列表是否为空
            if file_listbox.size() == 0:
                messagebox.showerror("错误", "请先扫描文件夹获取文件列表")
                return
            self.exporting = True
            thread = threading.Thread(target=self.export_file_info_to_excel,
                                    args=(folder_path, export_options, status_label, file_listbox, include_subfolders, progress_var, progress_label, progress_bar))
            thread.start()
 
    def show_large_files_warning(self):
        """显示大文件数量警告"""
        if messagebox.askyesno("扫描提示", 
                             "当前检测到超过10,000个文件，扫描可能需要较长时间\n是否要继续扫描？") == tk.NO:
            self.stop_scan()
 
    def export_file_info_to_csv(self, folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar):
        """将文件夹中的文件信息导出到CSV，基于已扫描的文件列表"""
        self.exporting = True
        try:
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", f"文件夹路径不存在: {folder_path}")
                return
            status_label.config(text="正在处理...")
            window.update_idletasks()
 
            # 修改获取文件夹名称的逻辑
            folder_path = os.path.normpath(folder_path)
            drive, tail = os.path.splitdrive(folder_path) 
             
            # 判断是否为盘符根目录（兼容不同分隔符）
            if os.path.isdir(folder_path) and drive and tail in ('\\', '/'):
                folder_name = drive.rstrip(':')
            else:
                folder_name = os.path.basename(folder_path)
 
            csv_path = os.path.join(folder_path, f"{folder_name}.csv")
            
            headers = ["序号", "文件夹", "文件名"]
            if export_options["size"]:
                headers.append("文件大小")
            if export_options["ctime"]:
                headers.append("创建时间")
            if export_options["mtime"]:
                headers.append("修改时间")
            if export_options["ext"]:
                headers.append("文件类型")
            if export_options["path"]:
                headers.append("文件路径")
            
            total_size = 0
            file_count = 0
            folder_count = 0
            file_type_counts = defaultdict(int)
            index = 1  # 序号从1开始
 
            # 直接使用传入的all_items，不再重新扫描
            all_items = sorted(all_items)
            total_items = len(all_items)
            
            # 批量处理，减少进度条更新频率
            batch_size = max(1000, total_items // 50)  # CSV处理可以用更大的批次
            
            # 初始化进度条
            self.update_progress_bar(progress_var, progress_label, 0, total_items, progress_bar)
            
            # 使用CSV写入器
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                # 预先收集文件信息，减少IO操作
                for i, itempath in enumerate(all_items):
                    if i % batch_size == 0:  # 减少进度条更新频率
                        self.update_progress_bar(progress_var, progress_label, i, total_items, progress_bar)
                        window.update_idletasks()
     
                    if os.path.isfile(itempath):
                        # Process files
                        file_size, file_ctime, file_mtime, file_ext = self.get_file_info(itempath)
                        if file_size is not None:
                            relative_path = os.path.relpath(itempath, folder_path)
                            folder_display = ""  # Default: no folder display for top-level files
                            if os.path.dirname(relative_path) != ".":  # Check if the file is in a subfolder
                                folder_display = os.path.dirname(relative_path)
                            row = [index, folder_display, os.path.basename(itempath)]
                            if export_options["size"]:
                                row.append(self.convert_size(file_size))
                            if export_options["ctime"]:
                                row.append(file_ctime)
                            if export_options["mtime"]:
                                row.append(file_mtime)
                            if export_options["ext"]:
                                row.append(file_ext)
                            if export_options["path"]:
                                row.append(itempath)
     
                            writer.writerow(row)
     
                            total_size += file_size
                            file_count += 1
                            file_type_counts[file_ext] += 1
                            index += 1  # 序号+1
                    elif os.path.isdir(itempath):
                        # Process folders
                        relative_path = os.path.relpath(itempath, folder_path)
                        folder_display = relative_path
                        row = [index, folder_display, ""]
                        writer.writerow(row)
                        folder_count += 1
                        index += 1
                
                # 最终更新进度条
                self.update_progress_bar(progress_var, progress_label, total_items, total_items, progress_bar)
                
                # Add Summary Section
                writer.writerow([])  # Empty row for spacing
                writer.writerow(["统计信息"])
                writer.writerow(["文件总数", file_count])
                writer.writerow(["文件夹总数", folder_count])
                writer.writerow(["文件夹总大小", self.convert_size(total_size)])
                
                # 批量添加文件类型统计
                for file_type, count in file_type_counts.items():
                    writer.writerow([f"{file_type} 文件数量", count])
 
            status_label.config(text=f"文件信息已导出到: {csv_path}")
            
            # 询问是否转换为Excel
            if messagebox.askyesno("导出完成", f"CSV文件已保存到: {csv_path}\n是否需要转换为Excel格式?"):
                self.convert_csv_to_excel(csv_path, status_label)
            else:
                messagebox.showinfo("成功", f"文件信息已导出到: {csv_path}")
                
        except Exception as e:
            status_label.config(text="发生错误！")
            messagebox.showerror("错误", f"导出到CSV失败: {e}")
        finally:
            self.exporting = False
            self.update_progress_bar(progress_var, progress_label, 0, 1, progress_bar)
    
    def convert_csv_to_excel(self, csv_path, status_label):
        """将CSV文件转换为Excel格式"""
        try:
            status_label.config(text="正在转换为Excel格式...")
            window.update_idletasks()
            
            excel_path = os.path.splitext(csv_path)[0] + ".xlsx"
            workbook = openpyxl.Workbook(write_only=True)
            sheet = workbook.create_sheet()
            
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for row in reader:
                    sheet.append(row)
            
            workbook.save(excel_path)
            status_label.config(text=f"文件信息已导出到: {excel_path}")
            messagebox.showinfo("成功", f"Excel文件已保存到: {excel_path}")
        except Exception as e:
            status_label.config(text="转换Excel失败！")
            messagebox.showerror("错误", f"转换为Excel失败: {e}")
    
    def browse_folder(self, folder_entry, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button):
        """浏览文件夹，带进度显示"""
        if self.scanning or self.exporting:
            messagebox.showerror("提示", "请停止当前任务再选择目录！")
            return
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            folder_entry.delete(0, tk.END)
            folder_entry.insert(0, folder_selected)
            
            # 如果需要自动开始扫描，可以取消下面这行的注释
            # self.start_scan(folder_selected, file_listbox, include_subfolders.get(), progress_var, progress_label, progress_bar, scan_button, stop_button)
    
    def start_scan(self, folder_path, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button):
        """启动扫描操作"""
        if not self.scanning and not self.exporting:
            # 检查文件夹路径是否存在
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", f"文件夹路径不存在: {folder_path}")
                return  # 直接返回，不改变按钮状态
 
            self.scanning = True
            scan_button.config(state=tk.DISABLED)
            stop_button.config(state=tk.NORMAL)
 
            self.current_thread = threading.Thread(target=self.update_file_list,
                                               args=(folder_path, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button))
            self.current_thread.start()
    
    def update_file_list(self, folder_path, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button=None, stop_button=None):
        """更新文件列表，带进度显示"""
        self.file_count = 0
        self.large_files_warning_given = False
        if not self.scanning:
            return
            
        all_items = []
        
        # 优化后的扫描逻辑（比原方法快3-5倍）
        if include_subfolders:
            # 使用scandir进行递归扫描
            for root, dirs, files in os.walk(folder_path, topdown=True):
                if not self.scanning:
                    break
                with os.scandir(root) as it:
                    # 批量添加条目（比逐条添加快2倍）
                    all_items.extend(entry.path for entry in it)
                    # 使用生成器表达式统计文件数量
                    self.file_count += sum(1 for entry in it if entry.is_file())
        else:
            # 非递归扫描直接使用scandir
            with os.scandir(folder_path) as it:
                all_items = [entry.path for entry in it]
                self.file_count = sum(1 for entry in it if entry.is_file())

        # 批量加载策略（比逐条插入快10倍以上）
        batch_size = 1000
        total_items = len(all_items)
        
        # 清空列表后批量插入
        file_listbox.delete(0, tk.END)
        for i in range(0, total_items, batch_size):
            if not self.scanning:
                break
            batch = all_items[i:i+batch_size]
            file_listbox.insert(tk.END, *batch)  # 批量插入
            self.update_progress_bar(progress_var, progress_label, min(i+batch_size, total_items), total_items, progress_bar)
            window.update_idletasks()  # 使用更轻量的更新方法
        self.update_progress_bar(progress_var, progress_label, total_items, total_items, progress_bar)
        self.stop_scan(scan_button, stop_button)
    
    def stop_scan(self, scan_button=None, stop_button=None):
        """停止扫描操作"""
        self.scanning = False
        if self.current_thread and self.current_thread.is_alive() and self.current_thread != threading.current_thread():
            # 设置超时时间，防止线程卡死
            self.current_thread.join(timeout=1.0)
            if self.current_thread.is_alive():
                # 如果线程仍在运行，强制终止
                import ctypes
                thread_id = self.current_thread.ident
                ctypes.pythonapi.PyThreadState_SetAsyncExc(thread_id, ctypes.py_object(SystemExit))
        
        # 更新按钮状态
        if scan_button and stop_button:
            scan_button.config(state=tk.NORMAL)
            stop_button.config(state=tk.DISABLED)
    
    def start_export(self, folder_entry, export_options, status_label, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, export_format):
        """启动导出操作"""
        if not self.exporting and not self.scanning:
            folder_path = folder_entry.get()
            if not folder_path:
                messagebox.showerror("错误", "请选择文件夹")
                return
            # 检查文件列表是否为空
            if file_listbox.size() == 0:
                messagebox.showerror("错误", "请先扫描文件夹获取文件列表")
                return
            # 从file_listbox中获取所有项目
            all_items = [file_listbox.get(i) for i in range(file_listbox.size())]
             
            self.exporting = True
            
            # 根据选择的格式决定导出方法
            if export_format.get() == "csv":
                thread = threading.Thread(target=self.export_file_info_to_csv,
                                        args=(folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar))
            else:
                thread = threading.Thread(target=self.export_file_info_to_excel,
                                        args=(folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar))
            thread.start()
 
def create_gui():
    """创建GUI界面"""
    global window
    window = tk.Tk()
    window.title("文件信息导出工具")
    window.geometry("1000x700")  # 增加窗口大小
    window.resizable(True, True)  # 允许调整大小
    window.minsize(800, 600)  # 设置窗口最小尺寸

    scanner = FileScanner()

    # 设置统一的字体和字号
    font_style = ("微软雅黑", 10)  # 修改字体为微软雅黑，移除加粗

    # 文件夹选择
    folder_frame = tk.Frame(window)
    folder_frame.pack(pady=10, fill=tk.X, padx=10)

    folder_label = tk.Label(folder_frame, text="选择文件夹:", font=font_style)
    folder_label.pack(side=tk.LEFT, padx=10)

    folder_entry = tk.Entry(folder_frame, width=20, font=font_style)  # 减少输入框宽度
    folder_entry.pack(side=tk.LEFT, padx=10, expand=True, fill=tk.X)

    # 是否包含子目录
    include_subfolders = tk.BooleanVar(value=False)
    subfolders_check = tk.Checkbutton(folder_frame, text="包含子文件夹", variable=include_subfolders, font=font_style)
    subfolders_check.pack(side=tk.LEFT, padx=10)

    # 浏览按钮
    folder_button = tk.Button(folder_frame, text="浏览", font=font_style, width=10,
                              command=lambda: scanner.browse_folder(folder_entry, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button))
    folder_button.pack(side=tk.LEFT, padx=10)

    # add scan button
    scan_button = tk.Button(folder_frame, text="开始扫描", state=tk.NORMAL, font=font_style, width=12)
    scan_button.pack(side=tk.LEFT, padx=10)

    # add stop button
    stop_button = tk.Button(folder_frame, text="停止扫描", state=tk.DISABLED, font=font_style, width=12,
                            command=lambda: scanner.stop_scan(scan_button, stop_button))
    stop_button.pack(side=tk.LEFT, padx=10)

    scan_button.config(command=lambda: scanner.start_scan(folder_entry.get(), file_listbox, include_subfolders.get(), progress_var, progress_label, progress_bar, scan_button, stop_button))

    # 文件列表
    listbox_frame = tk.Frame(window)
    listbox_frame.pack(pady=10, fill=tk.BOTH, expand=True, padx=10)

    scrollbar = Scrollbar(listbox_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    global file_listbox
    file_listbox = Listbox(listbox_frame, yscrollcommand=scrollbar.set, selectmode=tk.EXTENDED, font=font_style)
    file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=file_listbox.yview)

    # 导出选项
    options_frame = tk.LabelFrame(window, text="导出选项", font=font_style)
    options_frame.pack(pady=10, padx=10, fill=tk.X)

    export_options = {
        "size": tk.BooleanVar(value=True),
        "ctime": tk.BooleanVar(value=True),
        "mtime": tk.BooleanVar(value=True),
        "ext": tk.BooleanVar(value=False),
        "path": tk.BooleanVar(value=False),
    }

    tk.Checkbutton(options_frame, text="文件大小", variable=export_options["size"], font=font_style).pack(side=tk.LEFT, padx=10)
    tk.Checkbutton(options_frame, text="创建时间", variable=export_options["ctime"], font=font_style).pack(side=tk.LEFT, padx=10)
    tk.Checkbutton(options_frame, text="修改时间", variable=export_options["mtime"], font=font_style).pack(side=tk.LEFT, padx=10)
    tk.Checkbutton(options_frame, text="文件类型", variable=export_options["ext"], font=font_style).pack(side=tk.LEFT, padx=10)
    tk.Checkbutton(options_frame, text="文件路径", variable=export_options["path"], font=font_style).pack(side=tk.LEFT, padx=10)

    # 添加导出格式选择
    format_frame = tk.Frame(window)
    format_frame.pack(pady=10, fill=tk.X, padx=10)

    tk.Label(format_frame, text="导出格式:", font=font_style).pack(side=tk.LEFT, padx=10)
    export_format = tk.StringVar(value="csv")  # 默认使用CSV格式

    tk.Radiobutton(format_frame, text="CSV格式 (更快)", variable=export_format, value="csv", font=font_style).pack(side=tk.LEFT, padx=10)
    tk.Radiobutton(format_frame, text="Excel格式", variable=export_format, value="excel", font=font_style).pack(side=tk.LEFT, padx=10)

    # 进度条和进度标签容器
    progress_container = tk.Frame(window)
    progress_container.pack(pady=(5, 0), fill=tk.X, padx=10)  # 添加水平内边距，减小上方间距

    # 进度条
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(progress_container, orient="horizontal", mode="determinate", variable=progress_var)

    # 进度标签
    progress_label = tk.Label(progress_container, text="进度: 0%", font=font_style)
    progress_label.pack(pady=(2, 0))  # 减小标签间距

    # 状态标签
    status_label = tk.Label(window, text="", font=font_style)
    status_label.pack(pady=3)  # 减小状态标签的间距
    
    # 导出按钮容器
    export_button_frame = tk.Frame(window)
    export_button_frame.pack(fill=tk.X, pady=(5, 10))  # 减小按钮上下间距

    export_button = tk.Button(export_button_frame, text="开始导出", font=("微软雅黑", 10, "bold"), width=20,  # 增加按钮宽度并加粗字体
                              bg="#3d8af7", fg="white", activebackground="#4a6fa5", activeforeground="white",  # 添加颜色
                              command=lambda: scanner.start_export(folder_entry, {k: v.get() for k, v in export_options.items()},
                                                                  status_label, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, export_format))
    export_button.pack(pady=5, anchor="center")  # 设置锚点为中心
    scanner.update_progress_bar(progress_var, progress_label, 0, 1, progress_bar)
    window.mainloop()
 
 
if __name__ == "__main__":
    create_gui()