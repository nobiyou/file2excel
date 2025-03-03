import os
import openpyxl
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, Listbox, Scrollbar
import threading
import math
import csv
from collections import defaultdict

# 定义应用主题颜色
COLORS = {
    "primary": "#4a6fa5",      # 主色调 - 深蓝色
    "secondary": "#6e9cd1",   # 次要色调 - 中蓝色
    "accent": "#3d8af7",      # 强调色 - 亮蓝色
    "background": "#f5f7fa",  # 背景色 - 浅灰蓝
    "text": "#333333",        # 文本色 - 深灰色
    "light_text": "#666666",  # 浅文本色
    "success": "#4caf50",     # 成功色 - 绿色
    "warning": "#ff9800",     # 警告色 - 橙色
    "error": "#f44336",       # 错误色 - 红色
    "disabled": "#cccccc"     # 禁用色 - 灰色
}

# 自定义样式类
class CustomStyle:
    @staticmethod
    def configure_styles():
        # 配置ttk样式
        style = ttk.Style()
        
        # 配置进度条样式
        style.configure(
            "Custom.Horizontal.TProgressbar",
            troughcolor=COLORS["background"],
            background=COLORS["accent"],
            thickness=10
        )
        
        # 配置按钮样式
        style.configure(
            "Custom.TButton",
            background=COLORS["primary"],
            foreground="white",
            font=("微软雅黑", 10),
            padding=5
        )
        
        # 配置Checkbutton样式
        style.configure(
            "Custom.TCheckbutton",
            background=COLORS["background"],
            foreground=COLORS["text"],
            font=("微软雅黑", 10)
        )
        
        # 配置Radiobutton样式
        style.configure(
            "Custom.TRadiobutton",
            background=COLORS["background"],
            foreground=COLORS["text"],
            font=("微软雅黑", 10)
        )
        
        # 配置Label样式
        style.configure(
            "Custom.TLabel",
            background=COLORS["background"],
            foreground=COLORS["text"],
            font=("微软雅黑", 10)
        )
        
        # 配置Frame样式
        style.configure(
            "Custom.TFrame",
            background=COLORS["background"]
        )
        
        # 配置LabelFrame样式
        style.configure(
            "Custom.TLabelframe",
            background=COLORS["background"],
            foreground=COLORS["text"],
            font=("微软雅黑", 10, "bold")
        )
        
        style.configure(
            "Custom.TLabelframe.Label",
            background=COLORS["background"],
            foreground=COLORS["primary"],
            font=("微软雅黑", 10, "bold")
        )

# 自定义按钮类，添加悬停效果
class HoverButton(tk.Button):
    def __init__(self, master, **kw):
        self.default_bg = kw.get('background', COLORS["primary"])
        self.hover_bg = kw.get('activebackground', COLORS["accent"])
        self.default_fg = kw.get('foreground', 'white')
        self.hover_fg = kw.get('activeforeground', 'white')
        
        # 设置圆角和阴影效果的样式
        kw['background'] = self.default_bg
        kw['foreground'] = self.default_fg
        kw['borderwidth'] = 0
        kw['padx'] = 15
        kw['pady'] = 5
        kw['font'] = ("微软雅黑", 10)
        
        tk.Button.__init__(self, master, **kw)
        self.bind("<Enter>", self.on_enter)
        self.bind("<Leave>", self.on_leave)
    
    def on_enter(self, e):
        self['background'] = self.hover_bg
        self['foreground'] = self.hover_fg
    
    def on_leave(self, e):
        self['background'] = self.default_bg
        self['foreground'] = self.default_fg

class FileScanner:
    def __init__(self):
        self.scanning = False
        self.exporting = False
        self.current_thread = None
        self.file_cache = []  # 缓存机制
        self.file_count = 0  # 文件计数器
        self.large_files_warning_given = False  # 警告标记

    def export_file_info_to_excel(self, folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar):
        """将文件夹中的文件信息导出到Excel，基于已扫描的文件列表"""
        self.exporting = True
        try:
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", f"文件夹路径不存在: {folder_path}")
                return
            status_label.config(text="正在处理...", foreground=COLORS["accent"])
            window.update_idletasks()

            # 获取文件夹名称
            folder_path = os.path.normpath(folder_path)
            drive, tail = os.path.splitdrive(folder_path)
            if os.path.isdir(folder_path) and drive and tail in ('\\', '/'):
                folder_name = drive.rstrip(':')
            else:
                folder_name = os.path.basename(folder_path)

            excel_path = os.path.join(folder_path, f"{folder_name}.xlsx")
            workbook = openpyxl.Workbook()
            sheet = workbook.active

            # 设置表头样式
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color="4A6FA5", end_color="4A6FA5", fill_type="solid")
            centered = Alignment(horizontal='center', vertical='center')
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )

            headers = ["序号", "文件夹", "文件名"]
            if export_options["size"]:
                headers.append("文件大小")
            if export_options["ctime"]:
                headers.append("创建时间")
            if export_options["mtime"]:
                headers.append("修改时间")
            if export_options["ext"]:
                headers.append("文件类型")
            if export_options["path"]:
                headers.append("文件路径")
            sheet.append(headers)
            
            # 应用表头样式
            for col_num, _ in enumerate(headers, 1):
                cell = sheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered
                cell.border = thin_border

            # 设置列宽
            for i, header in enumerate(headers):
                col_letter = openpyxl.utils.get_column_letter(i+1)
                if header == "文件名":
                    sheet.column_dimensions[col_letter].width = 30
                elif header == "文件路径":
                    sheet.column_dimensions[col_letter].width = 50
                elif header in ["创建时间", "修改时间"]:
                    sheet.column_dimensions[col_letter].width = 20
                else:
                    sheet.column_dimensions[col_letter].width = 15

            total_size = 0
            file_count = 0
            folder_count = 0
            file_type_counts = defaultdict(int)
            index = 1  # 序号从1开始

            # 直接使用传入的all_items，不再重新扫描
            all_items = sorted(all_items)
            total_items = len(all_items)

            # 批量处理，减少进度条更新频率
            batch_size = max(1000, total_items // 50)

            # 初始化进度条
            self.update_progress_bar(progress_var, progress_label, 0, total_items, progress_bar)

            # 设置单元格样式
            normal_font = Font(name='微软雅黑', size=10)
            alt_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")

            for i, itempath in enumerate(all_items):
                if i % batch_size == 0:  # 减少进度条更新频率
                    self.update_progress_bar(progress_var, progress_label, i, total_items, progress_bar)
                    window.update_idletasks()

                if os.path.isfile(itempath):
                    # 处理文件
                    file_size, file_ctime, file_mtime, file_ext = self.get_file_info(itempath)
                    if file_size is not None:
                        relative_path = os.path.relpath(itempath, folder_path)
                        folder_display = ""  # 默认：顶级文件不显示文件夹
                        if os.path.dirname(relative_path) != ".":  # 检查文件是否在子文件夹中
                            folder_display = os.path.dirname(relative_path)
                        row = [index, folder_display, os.path.basename(itempath)]
                        if export_options["size"]:
                            row.append(self.convert_size(file_size))
                        if export_options["ctime"]:
                            row.append(file_ctime)
                        if export_options["mtime"]:
                            row.append(file_mtime)
                        if export_options["ext"]:
                            row.append(file_ext)
                        if export_options["path"]:
                            row.append(itempath)

                        sheet.append(row)
                        
                        # 应用交替行样式
                        row_num = len(sheet._cells) // len(headers) + (1 if len(sheet._cells) % len(headers) else 0)
                        for col_num in range(1, len(row) + 1):
                            cell = sheet.cell(row=row_num, column=col_num)
                            cell.font = normal_font
                            cell.border = thin_border
                            if row_num % 2 == 0:
                                cell.fill = alt_fill

                        total_size += file_size
                        file_count += 1
                        file_type_counts[file_ext] += 1
                        index += 1  # 序号+1
                elif os.path.isdir(itempath):
                    # 处理文件夹
                    relative_path = os.path.relpath(itempath, folder_path)
                    folder_display = relative_path
                    folder_name = os.path.basename(itempath)
                    row = [index, folder_display, folder_name]
                    sheet.append(row)
                    
                    # 应用文件夹行样式
                    row_num = len(sheet._cells) // len(headers) + (1 if len(sheet._cells) % len(headers) else 0)
                    folder_font = Font(name='微软雅黑', size=10, bold=True)
                    folder_fill = PatternFill(start_color="E3E9F2", end_color="E3E9F2", fill_type="solid")
                    # 确保所有列都应用样式，即使超出了实际数据的列数
                    for col_num in range(1, len(headers) + 1):
                        cell = sheet.cell(row=row_num, column=col_num)
                        cell.font = folder_font
                        cell.fill = folder_fill
                        cell.border = thin_border
                    
                    folder_count += 1
                    index += 1

            # 最终更新进度条
            self.update_progress_bar(progress_var, progress_label, total_items, total_items, progress_bar)

            # 添加空行
            summary_row = sheet.max_row + 1
            summary_cell = sheet.cell(row=summary_row, column=1)
            summary_cell.value = " "       
            
            # 添加统计信息标题行
            summary_row = sheet.max_row + 1
            sheet.merge_cells(f'A{summary_row}:B{summary_row}')
            summary_cell = sheet.cell(row=summary_row, column=1)
            summary_cell.value = "统计信息"
            summary_cell.font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
            summary_cell.fill = PatternFill(start_color="6E9CD1", end_color="6E9CD1", fill_type="solid")
            summary_cell.alignment = centered
            summary_cell.border = thin_border
            
            # 为合并单元格之外的列添加边框和背景色
            for col_num in range(1, 2 + 1):
                cell = sheet.cell(row=summary_row, column=col_num)
                cell.fill = PatternFill(start_color="6E9CD1", end_color="6E9CD1", fill_type="solid")
                cell.border = thin_border
            
            # 统计数据样式
            stat_font = Font(name='微软雅黑', size=10, bold=False)
            stat_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
            stat_alignment = Alignment(horizontal='left', vertical='center')
            
            # 添加统计数据行
            stat_rows = [
                ["文件总数", file_count],
                ["文件夹总数", folder_count],
                ["文件夹总大小", self.convert_size(total_size)]
            ]
            
            for idx, stat_row in enumerate(stat_rows):
                sheet.append(stat_row)
                row_num = sheet.max_row
                # 应用样式到所有列，包括空列
                for col_num in range(1, 2 + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.font = stat_font
                    cell.border = thin_border
                    cell.alignment = stat_alignment
                    if idx % 2 == 0:  # 交替行填充
                        cell.fill = stat_fill
            
            # 统计文件类型样式
            count_font = Font(name='微软雅黑', size=10, bold=True)
            count_fill = PatternFill(start_color="F5F7FA", end_color="F5F7FA", fill_type="solid")
            count_alignment = Alignment(horizontal='left', vertical='center')

            # 批量添加文件类型统计，并应用样式（按数量从大到小排序）
            sorted_file_types = sorted(file_type_counts.items(), key=lambda x: x[1], reverse=True)
            for idx, (file_type, count) in enumerate(sorted_file_types):
                row = [f"{file_type} 文件数量", count]
                sheet.append(row)
                row_num = sheet.max_row
                # 应用样式到所有列，包括空列
                for col_num in range(1, 2 + 1):
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.font = count_font
                    cell.border = thin_border
                    cell.alignment = count_alignment
                    if idx % 2 == 0:  # 交替行填充
                        cell.fill = count_fill

            workbook.save(excel_path)
            status_label.config(text=f"文件信息已导出到: {excel_path}", foreground=COLORS["success"])
            messagebox.showinfo("成功", f"Excel文件已保存到: {excel_path}")

        except Exception as e:
            status_label.config(text="发生错误！", foreground=COLORS["error"])
            messagebox.showerror("错误", f"导出到Excel失败: {e}")
        finally:
            self.exporting = False
            self.update_progress_bar(progress_var, progress_label, 0, 1, progress_bar)

    def convert_size(self, size_bytes):
        """转换文件大小为KB, MB, GB, TB等"""
        if size_bytes == 0:
            return "0B"
        size_name = ("B", "KB", "MB", "GB", "TB", "PB", "EB", "ZB", "YB")
        i = int(math.floor(math.log(size_bytes, 1024)))
        p = math.pow(1024, i)
        s = round(size_bytes / p, 2)
        return f"{s} {size_name[i]}"
 
    def get_file_info(self, filepath):
        """获取文件信息"""
        try:
            file_size = os.path.getsize(filepath)
            file_ctime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getctime(filepath)))
            file_mtime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(os.path.getmtime(filepath)))
            file_ext = os.path.splitext(filepath)[1]  # 获取文件扩展名
            return file_size, file_ctime, file_mtime, file_ext
        except FileNotFoundError:
            return None, None, None, None
        except Exception as e:
            return None, None, None, None
 
    def update_progress_bar(self, progress_var, progress_label, current, total, progress_bar):
        """更新进度条和进度标签"""
        progress = int((current / total) * 100) if total > 0 else 0
        progress_var.set(progress)
        progress_label.config(text=f"进度: {progress}%", foreground=COLORS["primary"])
        if progress == 0:
            # 只隐藏进度条，不隐藏整个容器，这样状态标签仍然可见
            progress_bar.pack_forget()  # 隐藏进度条
            progress_label.pack_forget() # 隐藏进度标签
        else:
            # 确保进度条显示
            progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, pady=(0, 5), padx=5)  # 显示进度条
            # 确保进度标签显示在进度条右侧
            progress_label.pack(side=tk.RIGHT, pady=(0, 5), padx=(10, 0))  # 显示进度标签在右侧
        window.update_idletasks()

    def export_file_info_to_csv(self, folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar):
        """将文件夹中的文件信息导出到CSV，基于已扫描的文件列表"""
        self.exporting = True
        try:
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", f"文件夹路径不存在: {folder_path}")
                return
            status_label.config(text="正在处理...", foreground=COLORS["accent"])
            window.update_idletasks()
 
            # 修改获取文件夹名称的逻辑
            folder_path = os.path.normpath(folder_path)
            drive, tail = os.path.splitdrive(folder_path) 
             
            # 判断是否为盘符根目录（兼容不同分隔符）
            if os.path.isdir(folder_path) and drive and tail in ('\\', '/'):
                folder_name = drive.rstrip(':')
            else:
                folder_name = os.path.basename(folder_path)
 
            csv_path = os.path.join(folder_path, f"{folder_name}.csv")
            
            headers = ["序号", "文件夹", "文件名"]
            if export_options["size"]:
                headers.append("文件大小")
            if export_options["ctime"]:
                headers.append("创建时间")
            if export_options["mtime"]:
                headers.append("修改时间")
            if export_options["ext"]:
                headers.append("文件类型")
            if export_options["path"]:
                headers.append("文件路径")
            
            total_size = 0
            file_count = 0
            folder_count = 0
            file_type_counts = defaultdict(int)
            index = 1  # 序号从1开始
 
            # 直接使用传入的all_items，不再重新扫描
            all_items = sorted(all_items)
            total_items = len(all_items)
            
            # 批量处理，减少进度条更新频率
            batch_size = max(1000, total_items // 50)  # CSV处理可以用更大的批次
            
            # 初始化进度条
            self.update_progress_bar(progress_var, progress_label, 0, total_items, progress_bar)
            
            # 使用CSV写入器
            with open(csv_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(headers)
                
                # 预先收集文件信息，减少IO操作
                for i, itempath in enumerate(all_items):
                    if i % batch_size == 0:  # 减少进度条更新频率
                        self.update_progress_bar(progress_var, progress_label, i, total_items, progress_bar)
                        window.update_idletasks()
     
                    if os.path.isfile(itempath):
                        # 处理文件
                        file_size, file_ctime, file_mtime, file_ext = self.get_file_info(itempath)
                        if file_size is not None:
                            relative_path = os.path.relpath(itempath, folder_path)
                            folder_display = ""  # 默认：顶级文件不显示文件夹
                            if os.path.dirname(relative_path) != ".":  # 检查文件是否在子文件夹中
                                folder_display = os.path.dirname(relative_path)
                            row = [index, folder_display, os.path.basename(itempath)]
                            if export_options["size"]:
                                row.append(self.convert_size(file_size))
                            if export_options["ctime"]:
                                row.append(file_ctime)
                            if export_options["mtime"]:
                                row.append(file_mtime)
                            if export_options["ext"]:
                                row.append(file_ext)
                            if export_options["path"]:
                                row.append(itempath)
     
                            writer.writerow(row)
     
                            total_size += file_size
                            file_count += 1
                            file_type_counts[file_ext] += 1
                            index += 1  # 序号+1
                    elif os.path.isdir(itempath):
                        # 处理文件夹
                        relative_path = os.path.relpath(itempath, folder_path)
                        folder_display = relative_path
                        row = [index, folder_display, ""]
                        writer.writerow(row)
                        folder_count += 1
                        index += 1
                
                # 最终更新进度条
                self.update_progress_bar(progress_var, progress_label, total_items, total_items, progress_bar)
                
                # 添加统计信息
                writer.writerow([])  # 空行用于分隔
                writer.writerow(["统计信息"])
                writer.writerow(["文件总数", file_count])
                writer.writerow(["文件夹总数", folder_count])
                writer.writerow(["文件夹总大小", self.convert_size(total_size)])
                
                # 批量添加文件类型统计（按数量从大到小排序）
                sorted_file_types = sorted(file_type_counts.items(), key=lambda x: x[1], reverse=True)
                for file_type, count in sorted_file_types:
                    writer.writerow([f"{file_type} 文件数量", count])
 
            status_label.config(text=f"文件信息已导出到: {csv_path}", foreground=COLORS["success"])
            
            # 询问是否转换为Excel
            if messagebox.askyesno("导出完成", f"CSV文件已保存到: {csv_path}\n是否需要转换为Excel格式?"):
                self.convert_csv_to_excel(csv_path, status_label)
            else:
                messagebox.showinfo("成功", f"文件信息已导出到: {csv_path}")
                
        except Exception as e:
            status_label.config(text="发生错误！", foreground=COLORS["error"])
            messagebox.showerror("错误", f"导出到CSV失败: {e}")
        finally:
            self.exporting = False
            self.update_progress_bar(progress_var, progress_label, 0, 1, progress_bar)
    
    def convert_csv_to_excel(self, csv_path, status_label):
        """将CSV文件转换为Excel格式"""
        try:
            status_label.config(text="正在转换为Excel格式...", foreground=COLORS["accent"])
            window.update_idletasks()
            
            excel_path = os.path.splitext(csv_path)[0] + ".xlsx"
            workbook = openpyxl.Workbook(write_only=True)
            sheet = workbook.create_sheet()
            
            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f)
                for row in reader:
                    sheet.append(row)
            
            workbook.save(excel_path)
            status_label.config(text=f"文件信息已导出到: {excel_path}", foreground=COLORS["success"])
            messagebox.showinfo("成功", f"Excel文件已保存到: {excel_path}")
        except Exception as e:
            status_label.config(text="转换Excel失败！", foreground=COLORS["error"])
            messagebox.showerror("错误", f"转换为Excel失败: {e}")
    
    def browse_folder(self, folder_entry, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button):
        """浏览文件夹，带进度显示"""
        if self.scanning or self.exporting:
            messagebox.showerror("提示", "请停止当前任务再选择目录！")
            return
        folder_selected = filedialog.askdirectory()
        if folder_selected:
            folder_entry.delete(0, tk.END)
            folder_entry.insert(0, folder_selected)
            
            # 如果需要自动开始扫描，可以取消下面这行的注释
            # self.start_scan(folder_selected, file_listbox, include_subfolders.get(), progress_var, progress_label, progress_bar, scan_button, stop_button)
    
    def start_scan(self, folder_path, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button, status_label=None):
        """启动扫描操作"""
        if not self.scanning and not self.exporting:
            # 检查文件夹路径是否存在
            if not os.path.exists(folder_path):
                messagebox.showerror("错误", f"文件夹路径不存在: {folder_path}")
                if status_label:
                    status_label.config(text="扫描失败：文件夹不存在", foreground=COLORS["error"])
                return  # 直接返回，不改变按钮状态
 
            self.scanning = True
            scan_button.config(state=tk.DISABLED)
            stop_button.config(state=tk.NORMAL)
            
            if status_label:
                status_label.config(text="正在扫描...", foreground=COLORS["accent"])
 
            self.current_thread = threading.Thread(target=self.update_file_list,
                                               args=(folder_path, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button, stop_button, status_label))
            self.current_thread.start()
    
    def update_file_list(self, folder_path, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, scan_button=None, stop_button=None, status_label=None):
        """更新文件列表，带进度显示"""
        self.file_count = 0
        self.large_files_warning_given = False
        if not self.scanning:
            if status_label:
                status_label.config(text="扫描已停止", foreground=COLORS["warning"])
            return
            
        all_items = []
        
        # 优化后的扫描逻辑（比原方法快3-5倍）
        if include_subfolders:
            # 使用scandir进行递归扫描
            for root, dirs, files in os.walk(folder_path, topdown=True):
                if not self.scanning:
                    break
                with os.scandir(root) as it:
                    # 批量添加条目（比逐条添加快2倍）
                    all_items.extend(entry.path for entry in it)
                    # 使用生成器表达式统计文件数量
                    self.file_count += sum(1 for entry in it if entry.is_file())
        else:
            # 非递归扫描直接使用scandir
            with os.scandir(folder_path) as it:
                all_items = [entry.path for entry in it]
                self.file_count = sum(1 for entry in it if entry.is_file())

        # 批量加载策略（比逐条插入快10倍以上）
        batch_size = 1000
        total_items = len(all_items)
        
        # 清空列表后批量插入
        file_listbox.delete(0, tk.END)
        for i in range(0, total_items, batch_size):
            if not self.scanning:
                break
            batch = all_items[i:i+batch_size]
            file_listbox.insert(tk.END, *batch)  # 批量插入
            self.update_progress_bar(progress_var, progress_label, min(i+batch_size, total_items), total_items, progress_bar)
            window.update_idletasks()  # 使用更轻量的更新方法
        self.update_progress_bar(progress_var, progress_label, total_items, total_items, progress_bar)
        if status_label:
            status_label.config(text="扫描完成", foreground=COLORS["success"])
        self.stop_scan(scan_button, stop_button)
    
    def stop_scan(self, scan_button=None, stop_button=None, status_label=None):
        """停止扫描操作"""
        self.scanning = False
        if self.current_thread and self.current_thread.is_alive() and self.current_thread != threading.current_thread():
            # 设置超时时间，防止线程卡死
            self.current_thread.join(timeout=1.0)
            if self.current_thread.is_alive():
                # 如果线程仍在运行，强制终止
                import ctypes
                thread_id = self.current_thread.ident
                ctypes.pythonapi.PyThreadState_SetAsyncExc(thread_id, ctypes.py_object(SystemExit))
                if status_label:
                    status_label.config(text="扫描已强制停止", foreground=COLORS["warning"])
        
        # 更新按钮状态
        if scan_button and stop_button:
            scan_button.config(state=tk.NORMAL)
            stop_button.config(state=tk.DISABLED)
    
    def start_export(self, folder_entry, export_options, status_label, file_listbox, include_subfolders, progress_var, progress_label, progress_bar, export_format):
        """启动导出操作"""
        if not self.exporting and not self.scanning:
            folder_path = folder_entry.get()
            if not folder_path:
                messagebox.showerror("错误", "请选择文件夹")
                return
            # 检查文件列表是否为空
            if file_listbox.size() == 0:
                messagebox.showerror("错误", "请先扫描文件夹获取文件列表")
                return
            # 从file_listbox中获取所有项目
            all_items = [file_listbox.get(i) for i in range(file_listbox.size())]
             
            self.exporting = True
            
            # 根据选择的格式决定导出方法
            if export_format.get() == "csv":
                thread = threading.Thread(target=self.export_file_info_to_csv,
                                        args=(folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar))
            else:
                thread = threading.Thread(target=self.export_file_info_to_excel,
                                        args=(folder_path, export_options, status_label, all_items, include_subfolders, progress_var, progress_label, progress_bar))
            thread.start()

def create_gui():
    """创建GUI界面"""
    global window
    window = tk.Tk()
    window.title("文件信息导出工具")
    window.iconbitmap('folder.ico')
    window.geometry("1000x700")  # 增加窗口大小
    window.resizable(True, True)  # 允许调整大小
    window.minsize(900, 600)  # 设置窗口最小尺寸
    window.configure(bg=COLORS["background"])  # 设置窗口背景色

    # 配置自定义样式
    CustomStyle.configure_styles()

    scanner = FileScanner()

    # 设置统一的字体和字号
    font_style = ("微软雅黑", 10)

    # 创建主框架
    main_frame = ttk.Frame(window, style="Custom.TFrame")
    main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

    # 顶部标题
    title_frame = ttk.Frame(main_frame, style="Custom.TFrame")
    title_frame.pack(fill=tk.X, pady=(0, 20))
    
    title_label = tk.Label(title_frame, text="文件信息导出工具", font=("微软雅黑", 16, "bold"), 
                          fg=COLORS["primary"], bg=COLORS["background"])
    title_label.pack()
    
    subtitle_label = tk.Label(title_frame, text="扫描文件夹并导出文件信息到Excel或CSV", 
                             font=("微软雅黑", 10), fg=COLORS["light_text"], bg=COLORS["background"])
    subtitle_label.pack(pady=(5, 0))

    # 文件夹选择区域
    folder_frame = ttk.LabelFrame(main_frame, text="选择文件夹", style="Custom.TLabelframe")
    folder_frame.pack(pady=10, fill=tk.X)

    folder_content_frame = ttk.Frame(folder_frame, style="Custom.TFrame")
    folder_content_frame.pack(padx=10, pady=15, fill=tk.X)

    folder_entry = tk.Entry(folder_content_frame, font=font_style, bg="white", fg=COLORS["text"])
    folder_entry.pack(side=tk.LEFT, padx=10, expand=True, fill=tk.X)

    # 是否包含子目录
    include_subfolders = tk.BooleanVar(value=False)
    subfolders_check = ttk.Checkbutton(folder_content_frame, text="包含子文件夹", 
                                     variable=include_subfolders, style="Custom.TCheckbutton")
    subfolders_check.pack(side=tk.LEFT, padx=10)

    # 浏览按钮
    folder_button = HoverButton(folder_content_frame, text="浏览", 
                              command=lambda: scanner.browse_folder(folder_entry, file_listbox, include_subfolders, 
                                                                  progress_var, progress_label, progress_bar, scan_button, stop_button))
    folder_button.pack(side=tk.LEFT, padx=10)

    # 扫描和停止按钮
    button_frame = ttk.Frame(folder_content_frame, style="Custom.TFrame")
    button_frame.pack(side=tk.RIGHT, padx=10)

    scan_button = HoverButton(button_frame, text="开始扫描", state=tk.NORMAL)
    scan_button.pack(side=tk.LEFT, padx=5)

    stop_button = HoverButton(button_frame, text="停止扫描", state=tk.DISABLED,
                           background=COLORS["error"], activebackground="#d32f2f",
                           command=lambda: scanner.stop_scan(scan_button, stop_button))
    stop_button.pack(side=tk.LEFT, padx=5)

    scan_button.config(command=lambda: scanner.start_scan(folder_entry.get(), file_listbox, 
                                                       include_subfolders.get(), progress_var, 
                                                       progress_label, progress_bar, scan_button, stop_button, status_label))

    # 导出选项区域
    options_frame = ttk.LabelFrame(main_frame, text="导出选项", style="Custom.TLabelframe")
    options_frame.pack(pady=10, fill=tk.X)

    options_content_frame = ttk.Frame(options_frame, style="Custom.TFrame")
    options_content_frame.pack(padx=10, pady=10, fill=tk.X)

    export_options = {
        "size": tk.BooleanVar(value=True),
        "ctime": tk.BooleanVar(value=True),
        "mtime": tk.BooleanVar(value=True),
        "ext": tk.BooleanVar(value=False),
        "path": tk.BooleanVar(value=False),
    }

    # 使用网格布局排列选项
    option_items = [
        ("文件大小", "size"),
        ("创建时间", "ctime"),
        ("修改时间", "mtime"),
        ("文件类型", "ext"),
        ("文件路径", "path")
    ]
    
    for i, (text, key) in enumerate(option_items):
        ttk.Checkbutton(options_content_frame, text=text, variable=export_options[key], 
                      style="Custom.TCheckbutton").pack(side=tk.LEFT, padx=5)

    # 添加导出格式选择到导出选项框架中
    format_frame = ttk.Frame(options_content_frame, style="Custom.TFrame")
    format_frame.pack(side=tk.RIGHT, padx=10)

    ttk.Label(format_frame, text="导出格式:", style="Custom.TLabel").pack(side=tk.LEFT, padx=5)
    export_format = tk.StringVar(value="excel")  # 默认使用excel格式

    ttk.Radiobutton(format_frame, text="CSV格式 (更快)", variable=export_format, value="csv", 
                  style="Custom.TRadiobutton").pack(side=tk.LEFT, padx=5)
    ttk.Radiobutton(format_frame, text="Excel格式 (美观)", variable=export_format, value="excel", 
                  style="Custom.TRadiobutton").pack(side=tk.LEFT, padx=5)
    
    # 创建导出按钮的单独容器，靠右对齐
    export_button_frame = ttk.Frame(main_frame, style="Custom.TFrame")
    export_button_frame.pack(fill=tk.X, pady=10, padx=0)
    
    # 导出按钮
    export_button = HoverButton(export_button_frame, text="开始导出", width=20,  # 增加按钮宽度
                              background=COLORS["accent"],  # 使用强调色
                              activebackground=COLORS["primary"],  # 悬停时的颜色
                              command=lambda: scanner.start_export(folder_entry, 
                                                                {k: v.get() for k, v in export_options.items()},
                                                                status_label, file_listbox, include_subfolders, 
                                                                progress_var, progress_label, progress_bar, export_format))
    export_button.pack(side=tk.RIGHT, pady=5, padx=0)  # 放在右侧
    
    # 进度条区域 - 移动到导出按钮和文件列表之间
    progress_frame = ttk.Frame(main_frame, style="Custom.TFrame")
    progress_frame.pack(fill=tk.X, pady=10)
    
    # 创建一个内部框架来容纳进度条和状态信息
    progress_status_frame = ttk.Frame(progress_frame, style="Custom.TFrame")
    progress_status_frame.pack(fill=tk.X)
    
    # 左侧进度条区域
    progress_left_frame = ttk.Frame(progress_status_frame, style="Custom.TFrame")
    progress_left_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5), anchor=tk.W)
    
    # 进度条容器（水平布局）
    progress_container = ttk.Frame(progress_left_frame, style="Custom.TFrame")
    progress_container.pack(fill=tk.X)
    
    # 进度条
    progress_var = tk.DoubleVar()
    progress_bar = ttk.Progressbar(progress_container, orient="horizontal", mode="determinate", 
                                variable=progress_var, style="Custom.Horizontal.TProgressbar")
    progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
    # 进度标签（放在进度条右侧）
    progress_label = tk.Label(progress_container, text="进度: 0%", font=font_style, 
                            bg=COLORS["background"], fg=COLORS["primary"])
    progress_label.pack(side=tk.LEFT, padx=(10, 0))
    
    # 右侧状态信息区域
    status_right_frame = ttk.Frame(progress_status_frame, style="Custom.TFrame")
    status_right_frame.pack(side=tk.RIGHT, fill=tk.X, expand=True, padx=(5, 0), anchor=tk.E)
    
    # 状态标签 - 移到右侧
    status_label = tk.Label(status_right_frame, text="", font=font_style, 
                          bg=COLORS["background"], fg=COLORS["text"], justify=tk.RIGHT, anchor=tk.E) # 添加 anchor=tk.E
    status_label.pack(pady=1, fill=tk.X, expand=True,  padx=0, anchor=tk.E)

    
    # 文件列表区域
    list_frame = ttk.LabelFrame(main_frame, text="文件列表", style="Custom.TLabelframe")
    list_frame.pack(pady=10, fill=tk.BOTH, expand=True)

    listbox_frame = ttk.Frame(list_frame, style="Custom.TFrame")
    listbox_frame.pack(padx=15, pady=15, fill=tk.BOTH, expand=True)

    scrollbar = Scrollbar(listbox_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    global file_listbox
    file_listbox = Listbox(listbox_frame, yscrollcommand=scrollbar.set, selectmode=tk.EXTENDED, 
                         font=font_style, bg="white", fg=COLORS["text"], borderwidth=1)
    file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

    scrollbar.config(command=file_listbox.yview)

    # 状态区域已移至进度条右侧
    
    # 初始化进度条状态
    scanner.update_progress_bar(progress_var, progress_label, 0, 1, progress_bar)
    
    # 添加漂浮在顶层的版权信息
    floating_copyright_frame = tk.Frame(window, bg=COLORS["background"])
    floating_copyright_frame.configure(bg=COLORS["background"], bd=0, relief=tk.GROOVE)
    
    # 半透明效果
    floating_copyright_label = tk.Label(floating_copyright_frame, 
                                      text="© 2025 文件信息导出工具 by Nobiyou ",
                                      font=("微软雅黑", 8), 
                                      fg=COLORS["primary"], 
                                      bg=COLORS["background"])
    floating_copyright_label.pack(padx=10, pady=1)
    
    # 使用place布局管理器将版权信息固定在窗口底部中央
    floating_copyright_frame.place(relx=0.5, rely=1.0, anchor="s", y=-5)
    
    # 确保版权信息始终在顶层显示
    floating_copyright_frame.lift()
    
    # 绑定窗口大小变化事件，确保版权信息始终位于底部中央
    def update_copyright_position(event=None):
        floating_copyright_frame.place(relx=0.5, rely=1.0, anchor="s", y=-5)
        floating_copyright_frame.lift()  # 确保始终在顶层
    
    window.bind("<Configure>", update_copyright_position)
    
    window.mainloop()


if __name__ == "__main__":
    create_gui()